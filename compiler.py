#!/usr/bin/env python3
import os
import sys
import json
import argparse
import openai

import openpyxl

def map_to_excel(json_data, excel_path, output_path=None):
    """
    Updates the existing Excel template (preserving formulas, formatting).
    If output_path is None, will overwrite excel_path in place.
    """
    # Accept single object or list of dicts
    if isinstance(json_data, dict) and "year" in json_data:
        data = [json_data]
    elif isinstance(json_data, list):
        data = json_data
    else:
        raise ValueError("Invalid JSON data format.")

    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    year_row_idx = 3  # Third row: years
    cat_col_idx = 2   # Column B: categories

    # Map years to column numbers
    years = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=year_row_idx, column=col).value
        if val is not None:
            key_str = str(val).strip()
            years[key_str] = col
            try:
                key_int = str(int(float(val)))
                years[key_int] = col
            except Exception:
                pass

    # Map categories to row numbers
    cats = {}
    for row in range(1, ws.max_row + 1):
        val = ws.cell(row=row, column=cat_col_idx).value
        if val is not None:
            key_str = str(val).strip()
            cats[key_str] = row

    print("Year columns found:", years)
    print("Categories found in template:", list(cats.keys()))

    # Category mapping for spreadsheet
    category_mapping = {
        "Revenue": "Revenue",
        "Cost of Goods Sold (COGS)": "Cost of Goods Sold (COGS)",
        "Operating Expenses": "Operating Expenses",
        "Other Income": "Other Income",
        "Plus Owner Salary+Super etc": "Plus Owner Salary+Super etc",
        "Plus Owner Benefits": "Plus Owner Benefits",
        "Total add backs": "Total add backs"
    }

    for entry in data:
        yr = str(entry["year"]).strip()
        if yr not in years:
            print(f"Year {yr} not found, skipping.")
            continue

        for category, value in entry.items():
            if category == "year":
                continue
            excel_cat = category_mapping.get(category)
            if not excel_cat:
                print(f"Category '{category}' not in mapping, skipping.")
                continue
            row = cats.get(excel_cat)
            col = years.get(yr)
            if not row or not col:
                print(f"Category '{category}' (mapped to '{excel_cat}') or year {yr} not found, skipping.")
                continue

            cell = ws.cell(row=row, column=col)
            if value is None:
                continue
            if value == 0 and cell.value not in (None, ""):
                continue
            cell.value = value

    save_path = output_path if output_path else excel_path
    wb.save(save_path)
    print(f"Saved to {save_path}")

def extract_with_gpt(extracted_data: dict, user_prompt: str, api_key: str) -> dict:
    client = openai.OpenAI(api_key=api_key)

    # Categories must match the spreadsheet template labels exactly:
    fixed_categories_example = category_mapping = {
    "Revenue": "Revenue",
    "Cost of Goods Sold (COGS)": "Cost of Goods Sold (COGS)",
    "Operating Expenses": "Less Operating Expenses",
    "Other Income": "Other Income",
    "Owner Salary+Super": "Plus Owner Salary+Super etc",
    "Plus Owner Salary+Super etc": "Plus Owner Salary+Super etc",
    "Owner Benefits": "Plus Owner Benefits",
    "Plus Owner Benefits": "Plus Owner Benefits",
    "Total add backs": "Total add backs"
}

    system_content = (
        "You are a financial data extractor. "
        "Given raw financial data in various forms, map all values into the following fixed categories exactly as named: \n"
        + json.dumps(fixed_categories_example, indent=2) + "\n"
        "Return ONLY a JSON object or list of objects containing these categories and their numeric values. "
        "Include the year as the 'year' key. "
        "Do not include any percentages or fields not listed. "
        "If a category is missing in the input, set its value to 0."
    )

    print("\n--- INPUT SCHEMA TO GPT (extracted_data) ---")
    print(json.dumps(extracted_data, indent=2))
    print("\n--- SYSTEM PROMPT (Categories) ---")
    print(json.dumps(fixed_categories_example, indent=2))
    print("\n--- USER PROMPT ---")
    print(user_prompt)
    print("\n--- SENDING TO GPT... ---\n")

    messages = [
        {"role": "system", "content": system_content},
        {"role": "user", "content": json.dumps(extracted_data)},
        {"role": "user", "content": user_prompt}
    ]

    resp = client.chat.completions.create(
        model="gpt-4",
        messages=messages,
        temperature=0
    )
    body = resp.choices[0].message.content.strip()
    print("--- RAW GPT RESPONSE ---")
    print(body)
    # Robust JSON handling
    try:
        values = json.loads(body)
    except Exception as e:
        print(f"Error parsing GPT output as JSON: {e}")
        print("GPT output was:")
        print(body)
        sys.exit(1)
    return values

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--data", required=True, help="Path to raw-extracted JSON file or '-' for stdin")
    parser.add_argument("--prompt", required=True, help="User's GPT prompt")
    parser.add_argument("--key", required=True, help="OpenAI API key")
    parser.add_argument("--template", help="Excel template file (to auto-fill with GPT output)")
    parser.add_argument("--output", help="Output Excel file (leave blank to overwrite)")
    args = parser.parse_args()

    # Support reading input from stdin if --data is '-'
    if args.data == "-":
        extracted = json.load(sys.stdin)
    else:
        with open(args.data) as f:
            extracted = json.load(f)

    structured_data = extract_with_gpt(extracted, args.prompt, args.key)
    print("\n--- FINAL STRUCTURED OUTPUT ---")
    print(json.dumps(structured_data, indent=2))

    # Save the GPT output to "GPT output.json" in the repo (current working) directory
    out_path = os.path.join(os.getcwd(), "GPT output.json")
    with open(out_path, "w") as f:
        json.dump(structured_data, f, indent=2)
    print(f"GPT output saved to {out_path}")

    # If Excel template specified, update the Excel with GPT results
    if args.template:
        map_to_excel(structured_data, args.template, args.output)