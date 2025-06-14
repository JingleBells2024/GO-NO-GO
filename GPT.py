#!/usr/bin/env python3
import os
import sys
import json
import openai
import argparse
from openpyxl import load_workbook

def extract_with_gpt(extracted_data: dict, user_prompt: str, api_key: str) -> dict:
    """
    Calls OpenAI to turn raw extracted_data + user_prompt 
    into a clean JSON of field→value.
    """
    openai.api_key = api_key
    messages = [
        {"role": "system",
         "content": "You are a financial data extractor. "
                    "Given raw P&L data, return ONLY a JSON object "
                    "mapping each field name to its numeric value."},
        {"role": "user", "content": json.dumps(extracted_data)},
        {"role": "user", "content": user_prompt}
    ]
    resp = openai.ChatCompletion.create(
        model="gpt-4",
        messages=messages,
        temperature=0
    )
    body = resp.choices[0].message.content.strip()
    print("\nGPT Response:\n", body)  # For dev visibility
    return json.loads(body)

def fill_template(template_path: str,
                  output_path: str,
                  values: dict,
                  cell_map: dict):
    """
    Writes each value into its mapped cell in the template,
    then saves to output_path.
    """
    wb = load_workbook(template_path)
    ws = wb.active

    for field, cell in cell_map.items():
        if field in values:
            ws[cell] = values[field]

    wb.save(output_path)

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True, help="Path to .xlsx template")
    parser.add_argument("--data",     required=True, help="Path to raw-extracted JSON file")
    parser.add_argument("--prompt",   required=True, help="The user’s GPT prompt")
    parser.add_argument("--year",     type=int, required=True, help="Year column to fill")
    parser.add_argument("--key",      required=True, help="OpenAI API key")
    args = parser.parse_args()

    # Load extracted JSON data
    with open(args.data, 'r') as f:
        extracted = json.load(f)

    # Call GPT
    values = extract_with_gpt(extracted, args.prompt, args.key)

    # Map values to cells
    maps = {
        2023: {"Revenue":"D10","COGS":"D11","NetProfit":"D12"},
        2024: {"Revenue":"E10","COGS":"E11","NetProfit":"E12"},
        2025: {"Revenue":"F10","COGS":"F11","NetProfit":"F12"},
    }
    cell_map = maps.get(args.year, {})

    # Fill and save
    output_path = os.path.splitext(args.template)[0] + f"_filled_{args.year}.xlsx"
    fill_template(args.template, output_path, values, cell_map)
    print(f"\n✅ Finished! Output saved to: {output_path}")