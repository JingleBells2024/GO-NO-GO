# gui.py
import sys
import os
import json
import subprocess
import shlex
import re
from PyQt5.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QPushButton,
    QTextEdit, QFileDialog, QMessageBox, QLabel, QLineEdit
)
from openpyxl import load_workbook

try:
    from PyPDF2 import PdfReader
    PDF_SUPPORTED = True
except ImportError:
    PDF_SUPPORTED = False

class FinancialAnalysis(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('Financial Analysis')
        self.resize(600, 550)

        self.fin_file = None
        self.tpl_file = None
        self.api_key = None

        layout = QVBoxLayout()

        # Upload Financial File
        btn1 = QPushButton('Upload Financial PDF/Excel')
        btn1.clicked.connect(self.upload_financial)
        layout.addWidget(btn1)
        self.fin_label = QLabel('No financial file selected')
        layout.addWidget(self.fin_label)

        # Upload Excel Template
        btn2 = QPushButton('Upload Excel Template')
        btn2.clicked.connect(self.upload_template)
        layout.addWidget(btn2)
        self.tpl_label = QLabel('No template selected')
        layout.addWidget(self.tpl_label)

        # API Key input
        api_label = QLabel('Enter API Key:')
        layout.addWidget(api_label)
        self.api_input = QLineEdit()
        self.api_input.setEchoMode(QLineEdit.Password)
        self.api_input.setPlaceholderText('sk-...')
        layout.addWidget(self.api_input)

        # Prompt input
        self.text_edit = QTextEdit()
        self.text_edit.setPlaceholderText('Enter prompt for ChatGPT here…')
        layout.addWidget(self.text_edit)

        # Submit
        submit_btn = QPushButton('Submit')
        submit_btn.clicked.connect(self.submit)
        layout.addWidget(submit_btn)

        self.setLayout(layout)

    def upload_financial(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            'Choose Financial File',
            filter='PDF Files (*.pdf);;Excel Files (*.xlsx *.xlsm)'
        )
        if path:
            self.fin_file = path
            self.fin_label.setText(os.path.basename(path))
            self.fin_label.setStyleSheet('color: green')
        else:
            self.fin_file = None
            self.fin_label.setText('No financial file selected')
            self.fin_label.setStyleSheet('color: red')

    def upload_template(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            'Choose Excel Template',
            filter='Excel Files (*.xlsx *.xlsm)'
        )
        if path:
            self.tpl_file = path
            self.tpl_label.setText(os.path.basename(path))
            self.tpl_label.setStyleSheet('color: green')
        else:
            self.tpl_file = None
            self.tpl_label.setText('No template selected')
            self.tpl_label.setStyleSheet('color: red')

    def extract_excel(self, path):
        wb = load_workbook(path, data_only=True)
        out = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            out[sheet] = [list(row) for row in ws.iter_rows(values_only=True)]
        return out

    def extract_pdf(self, path):
        if not PDF_SUPPORTED:
            QMessageBox.critical(self, 'Error', 'Install PyPDF2 to extract PDF files.')
            return {}
        reader = PdfReader(path)
        pages = [p.extract_text() or '' for p in reader.pages]
        return {'pages': pages}

    def submit(self):
        prompt = self.text_edit.toPlainText().strip()
        self.api_key = self.api_input.text().strip()
        # Validate inputs
        if not self.api_key:
            QMessageBox.warning(self, 'Error', 'Please enter your API key.')
            return
        if not self.fin_file:
            QMessageBox.warning(self, 'Error', 'Please upload a financial file.')
            return
        if not self.tpl_file:
            QMessageBox.warning(self, 'Error', 'Please upload an Excel template.')
            return
        if not prompt:
            QMessageBox.warning(self, 'Error', 'Please enter a prompt.')
            return

        # Extract raw data in-process
        ext = os.path.splitext(self.fin_file)[1].lower()
        if ext in ('.xlsx', '.xlsm'):
            extracted_data = self.extract_excel(self.fin_file)
        elif ext == '.pdf':
            extracted_data = self.extract_pdf(self.fin_file)
        else:
            QMessageBox.warning(self, 'Error', 'Unsupported file type for extraction.')
            return

        # Save extracted JSON
        extracted_file = os.path.abspath('extracted.json')
        with open(extracted_file, 'w') as f:
            json.dump(extracted_data, f, indent=2)

        # Parse year from prompt
        m = re.search(r'20\d{2}', prompt)
        if not m:
            QMessageBox.warning(self, 'Error', 'Could not find a 4-digit year in your prompt.')
            return
        year = m.group(0)

        # Launch AI processor (GPT.py) in new Terminal
        ai_script = os.path.join(os.path.dirname(__file__), 'GPT.py')
        cmd = (
            f'{shlex.quote(sys.executable)} {shlex.quote(ai_script)} '
            f'--template {shlex.quote(self.tpl_file)} '
            f'--data {shlex.quote(extracted_file)} '
            f'--prompt {shlex.quote(prompt)} '
            f'--year {year} '
            f'--key {shlex.quote(self.api_input.text().strip())}'
        )
        # two-step AppleScript invocation to avoid quoting issues
        apple_args = [
            'osascript',
            '-e', 'tell application "Terminal" to activate',
            '-e', f'tell application "Terminal" to do script "{cmd}" in front window'
        ]
        subprocess.Popen(apple_args)

        QMessageBox.information(self, 'Started', 'Extraction done; AI script launched.')

if __name__ == '__main__':
    app = QApplication(sys.argv)
    win = FinancialAnalysis()
    win.show()
    sys.exit(app.exec_())


# GPT.py
import os
import json
import openai
from openpyxl import load_workbook


def extract_with_gpt(extracted_data: dict, user_prompt: str, api_key: str) -> dict:
    openai.api_key = api_key
    messages = [
        {"role": "system", "content": (
            "You are a financial data extractor."
            " Given raw P&L data, return ONLY a JSON object"
            " mapping each field name to its numeric value."
        )},
        {"role": "user", "content": json.dumps(extracted_data)},
        {"role": "user", "content": user_prompt}
    ]
    resp = openai.ChatCompletion.create(
        model="gpt-4",
        messages=messages,
        temperature=0
    )
    body = resp.choices[0].message.content.strip()
    values = json.loads(body)
    # print JSON for verification
    print(json.dumps(values, indent=2))
    return values


def fill_template(template_path: str,
                  output_path: str,
                  values: dict,
                  cell_map: dict):
    wb = load_workbook(template_path)
    ws = wb.active
    for field, cell in cell_map.items():
        if field in values:
            ws[cell] = values[field]
    wb.save(output_path)


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("--template", required=True)
    parser.add_argument("--data",     required=True)
    parser.add_argument("--prompt",   required=True)
    parser.add_argument("--year",     type=int, required=True)
    parser.add_argument("--key",      required=True)
    args = parser.parse_args()

    extracted = json.load(open(args.data))
    values = extract_with_gpt(extracted, args.prompt, args.key)

    # define cell maps per year
    maps = {
        2023: {"Revenue":"D10","COGS":"D11","NetProfit":"D12"},
        2024: {"Revenue":"E10","COGS":"E11","NetProfit":"E12"},
        2025: {"Revenue":"F10","COGS":"F11","NetProfit":"F12"},
    }
    cell_map = maps.get(args.year, {})
    out = os.path.splitext(args.template)[0] + f"_filled_{args.year}.xlsx"
    fill_template(args.template, out, values, cell_map)
    print(f"Finished! Output saved to {out}")