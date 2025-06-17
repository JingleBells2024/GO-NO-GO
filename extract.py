#!/usr/bin/env python3
import sys, os, argparse, json, re
from openpyxl import load_workbook

# PDF support (needs PyPDF2)
try:
    from PyPDF2 import PdfReader
    _PDF_OK = True
except ImportError:
    _PDF_OK = False

# All categories, in order, matching your compiler/template
CATEGORIES = [
    "year",
    "Revenue",
    "Cost of Goods Sold (COGS)",
    "Operating Expenses",
    "Taxes",
    "Depreciation & Amortization",
    "Plus Interest",
    "Owner Salary+Super",
    "Owner Benefits",
    "Manager Salary",
    "Investor Salary",
    "One off Revenue Adjustments",
    "One off Expenses Adjustments",
    "Other Adjustments 1",
    "Other Adjustments 2",
    "Assets",
    "Liabilities",
    "Equity",
    "Other Income",
    "Total add backs"
]

def extract_excel(path):
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row and "Revenue" in [str(cell) for cell in row]:
            header_row = i
            break
    if header_row is None:
        sys.exit("Could not find header row in Excel file.")

    # Read all rows after header, map by year if possible
    data = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        year = None
        entry = {cat: 0 for cat in CATEGORIES}
        for j, value in enumerate(row):
            if value is None:
                continue
            colname = ws.cell(row=header_row, column=j+1).value
            if colname is not None:
                cat = colname.strip()
                if cat in CATEGORIES:
                    if cat == "year":
                        entry[cat] = int(str(value).strip())
                        year = entry[cat]
                    else:
                        entry[cat] = value
        if year:
            entry["year"] = year
            data.append(entry)
    return data

def extract_pdf(path):
    if not _PDF_OK:
        sys.exit("Install PyPDF2 (`pip install PyPDF2`) to extract PDF")
    reader = PdfReader(path)
    all_text = "\n".join([p.extract_text() or "" for p in reader.pages])
    return parse_financial_text(all_text)

def parse_financial_text(text):
    # Extract each year's section
    year_blocks = re.split(r"\b(20\d{2})(?:\s*\(.*?\))?\s*\n", text)
    records = []
    # year_blocks is like: ["", "2022", "<block>", "2023", "<block>", ...]
    for i in range(1, len(year_blocks), 2):
        year = int(year_blocks[i])
        block = year_blocks[i + 1]
        record = {cat: 0 for cat in CATEGORIES}
        record["year"] = year
        for cat in CATEGORIES[1:]:  # skip year itself
            # Search for "Label: value" in the block
            pattern = re.compile(rf"{re.escape(cat)}[:：]\s*([\d,\.]+)", re.IGNORECASE)
            m = pattern.search(block)
            if m:
                val = m.group(1).replace(",", "")
                try:
                    record[cat] = float(val)
                except Exception:
                    record[cat] = 0
        records.append(record)
    return records

def main():
    parser = argparse.ArgumentParser(description="Extract financial data from PDF/Excel to structured JSON.")
    parser.add_argument("file", help="Path to .pdf or .xlsx file")
    parser.add_argument("-o", "--output", help="Output JSON file", default="extracted_data.json")
    args = parser.parse_args()

    if not os.path.isfile(args.file):
        sys.exit(f"File not found: {args.file}")

    ext = os.path.splitext(args.file)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        data = extract_excel(args.file)
    elif ext == ".pdf":
        data = extract_pdf(args.file)
    else:
        sys.exit("Unsupported file type. Use .pdf or .xlsx")

    # Write output as valid JSON array
    with open(args.output, "w") as f:
        json.dump(data, f, indent=2)
    print(f"Extracted structured data saved to {args.output}")
    print(json.dumps(data, indent=2))

if __name__ == "__main__":
    main()